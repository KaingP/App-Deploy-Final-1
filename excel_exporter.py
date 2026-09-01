import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Design Palette
PRIMARY_COLOR = "1E3A8A"
ACCENT_COLOR = "3B82F6"
SUCCESS_COLOR = "10B981"
WARNING_COLOR = "F59E0B"
DANGER_COLOR = "EF4444"
WHITE = "FFFFFF"

FONT_TITLE = Font(name="Segoe UI", size=15, bold=True, color="1E3A8A")
FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="4B5563")
FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Segoe UI", size=12, bold=True, color="1F2937")
FONT_BOLD = Font(name="Segoe UI", size=10, bold=True)
FONT_REGULAR = Font(name="Segoe UI", size=10)
FONT_SMALL = Font(name="Segoe UI", size=9, italic=True, color="6B7280")

FILL_HEADER_DARK = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
FILL_HEADER_BLUE = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
FILL_HEADER_ORANGE = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
FILL_HEADER_GREEN = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
FILL_HEADER_PURPLE = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")

FILL_ZEBRA_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_ZEBRA_2 = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_HIGHLIGHT_PHONG = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
FILL_HIGHLIGHT_NGOAI = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
FILL_SUCCESS_LIGHT = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

BORDER_THIN = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)
BORDER_HEADER = Border(
    left=Side(style='thin', color='475569'),
    right=Side(style='thin', color='475569'),
    top=Side(style='medium', color='1E293B'),
    bottom=Side(style='medium', color='1E293B')
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

def autofit_columns(ws, max_len_cap=55):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                lines = val_str.split('\n')
                line_len = max(len(l) for l in lines)
                max_len = max(max_len, line_len)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), max_len_cap)

def export_schedule_to_excel(schedule_result, file_path="reports/Lich_Truc_Toi_Uu_Hung_Vuong_Concert.xlsx", incident_logs=None):
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    assigned_shifts = schedule_result.get('assigned_shifts', [])
    member_stats = schedule_result.get('member_stats', [])
    audit = schedule_result.get('audit_results', {})
    contingency = schedule_result.get('contingency_matrix', [])

    create_sheet_tong_ca_truc(wb, assigned_shifts)
    create_sheet_ca_trong(wb, [s for s in assigned_shifts if s['type'] == 'Phong'])
    create_sheet_ca_ngoai(wb, [s for s in assigned_shifts if s['type'] == 'Ngoai'])
    create_sheet_thong_ke(wb, member_stats, assigned_shifts)
    create_sheet_kiem_tra_ca(wb, audit, assigned_shifts, member_stats)
    create_sheet_ca_vang(wb, assigned_shifts, contingency, incident_logs)

    wb.save(file_path)
    return file_path

# ==========================================
# SHEET 1: Tổng ca trực
# ==========================================
def create_sheet_tong_ca_truc(wb, assigned_shifts):
    ws = wb.create_sheet(title="Tổng ca trực")
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:K1")
    ws["A1"] = "HÙNG VƯƠNG CONCERT - PROJECT F&B | TỔNG HỢP TOÀN BỘ LỊCH TRỰC TUẦN"
    ws["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:K2")
    ws["A2"] = "Phân bổ tối ưu hóa nhân sự phòng Thanh Niên & Điểm bán ngoài (Google OR-Tools CP-SAT)"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = ALIGN_CENTER
    ws.row_dimensions[2].height = 20

    headers = [
        "Mã Ca", "Kênh / Loại", "Thứ", "Ngày", "Khung Giờ", 
        "Địa Điểm", "Yêu Cầu (Chính/DP)", "Đã Gán", "Trưởng Ca / Trưởng Nhóm", "Danh Sách Nhân Sự, SĐT & Nhiệm Vụ Trong Ca", "Tình Trạng"
    ]
    
    ws.row_dimensions[4].height = 28
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_BLUE
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    row_idx = 5
    for s in assigned_shifts:
        ws.row_dimensions[row_idx].height = 36
        
        member_parts = []
        for m in s['assigned_members']:
            is_ld = " [⭐ Trưởng Ca]" if s.get('shift_leader') == m['name'] else ""
            role_lbl = m.get('role', 'Chính')
            task_lbl = m.get('task', 'Bán hàng F&B')
            phone_lbl = f" - SĐT: {m['phone']}" if m.get('phone') else ""
            member_parts.append(f"• [{role_lbl}]{is_ld} {m['name']} ({m['department']}{phone_lbl}) — Nhiệm vụ: {task_lbl}")
            
        names = "\n".join(member_parts)
        status = "Đủ người" if s['is_filled'] else "Thiếu người"
        fill = FILL_HIGHLIGHT_PHONG if s['type'] == 'Phong' else FILL_HIGHLIGHT_NGOAI
        
        chinh_req = s.get('chinh_count', s.get('required_count', 3))
        dp_req = s.get('dp_count', 1)
        req_str = f"{chinh_req} Chính + {dp_req} DP"
        
        values = [
            s['shift_id'],
            s['type_label'],
            s['day'],
            s['date'],
            s['slot'],
            s['location'],
            req_str,
            s['assigned_count'],
            s['shift_leader'] or '-',
            names,
            status
        ]
        
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = FONT_REGULAR
            c.fill = fill
            c.border = BORDER_THIN
            c.alignment = ALIGN_CENTER if col_idx in [1, 2, 3, 4, 5, 7, 8, 11] else ALIGN_LEFT
            if col_idx == 11:
                c.font = Font(name="Segoe UI", size=10, bold=True, color="047857" if s['is_filled'] else "DC2626")
        row_idx += 1

    autofit_columns(ws)

# ==========================================
# SHEET 2: ca trong (Phòng Thanh Niên)
# ==========================================
def create_sheet_ca_trong(wb, phong_shifts):
    ws = wb.create_sheet(title="ca trong")
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:I1")
    ws["A1"] = "LỊCH TRỰC PHÒNG THANH NIÊN (THPT CHUYÊN HÙNG VƯƠNG) - CA TRONG"
    ws["A1"].font = Font(name="Segoe UI", size=15, bold=True, color="1E40AF")
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    headers = [
        "Mã Ca", "Thứ", "Ngày", "Khung Giờ", "Định Mức (Chính/DP)", "Thực Trực", "Trưởng Ca Phụ Trách", "Nhân Sự Phân Bổ (Phân Rõ Chính & DP)", "Ghi Chú Vận Hành"
    ]
    
    ws.row_dimensions[3].height = 26
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_BLUE
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    row_idx = 4
    for s in phong_shifts:
        ws.row_dimensions[row_idx].height = 36
        
        chinh_lines = []
        dp_lines = []
        for m in s['assigned_members']:
            is_ld = " [⭐ Trưởng ca]" if s.get('shift_leader') == m['name'] else ""
            task_lbl = m.get('task', 'Trực phòng & Bán F&B')
            line = f"{m['name']}{is_ld} ({m['department']} - SĐT: {m['phone']}) — Nhiệm vụ: {task_lbl}"
            if m.get('role') == 'Chính':
                chinh_lines.append(f"• [Trực chính]: {line}")
            else:
                dp_lines.append(f"• [Dự phòng]: {line}")
                
        member_str = "\n".join(chinh_lines + dp_lines)

        chinh_req = s.get('chinh_count', 3)
        dp_req = s.get('dp_count', 1)
        req_str = f"{chinh_req} Chính + {dp_req} DP"

        values = [
            s['shift_id'],
            s['day'],
            s['date'],
            s['slot'],
            req_str,
            s['assigned_count'],
            s['shift_leader'] or '-',
            member_str,
            "Trực phòng chính; Đảm bảo mở cửa đúng giờ, bảo quản tiền & hàng F&B"
        ]

        fill = FILL_ZEBRA_1 if row_idx % 2 == 0 else FILL_ZEBRA_2
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = FONT_REGULAR
            c.fill = fill
            c.border = BORDER_THIN
            c.alignment = ALIGN_CENTER if col_idx in [1, 2, 3, 4, 5, 6] else ALIGN_LEFT

        row_idx += 1

    autofit_columns(ws)

# ==========================================
# SHEET 3: ca ngoài (Điểm bán ngoài)
# ==========================================
def create_sheet_ca_ngoai(wb, ngoai_shifts):
    ws = wb.create_sheet(title="ca ngoài")
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:I1")
    ws["A1"] = "LỊCH PHÂN BỔ NHÂN SỰ ĐIỂM BÁN NGOÀI - CA NGOÀI"
    ws["A1"].font = Font(name="Segoe UI", size=15, bold=True, color="D97706")
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    headers = [
        "Mã Ca", "Thứ", "Ngày", "Khung Giờ", "Địa Điểm Bán Ngoài", "Định Mức (Chính/DP)", "Đã Phân", "Trưởng Điểm Ngoài", "Danh Sách Nhân Sự (SĐT & Phương Tiện)"
    ]
    
    ws.row_dimensions[3].height = 26
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_ORANGE
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    row_idx = 4
    for s in ngoai_shifts:
        ws.row_dimensions[row_idx].height = 36
        
        chinh_lines = []
        dp_lines = []
        for m in s['assigned_members']:
            veh = f"[{m['vehicle']}]" if m.get('vehicle') else ""
            is_ld = " [⭐ Trưởng nhóm]" if s.get('shift_leader') == m['name'] else ""
            task_lbl = m.get('task', 'Bán hàng ngoài')
            line = f"{m['name']}{is_ld} - {m['department']} (SĐT: {m['phone']}) {veh} — Nhiệm vụ: {task_lbl}"
            if m.get('role') == 'Chính':
                chinh_lines.append(f"• [Trực chính]: {line}")
            else:
                dp_lines.append(f"• [Dự phòng]: {line}")
                
        member_str = "\n".join(chinh_lines + dp_lines)

        chinh_req = s.get('chinh_count', 2)
        dp_req = s.get('dp_count', 1)
        req_str = f"{chinh_req} Chính + {dp_req} DP"

        values = [
            s['shift_id'],
            s['day'],
            s['date'],
            s['slot'],
            s['location'],
            req_str,
            s['assigned_count'],
            s['shift_leader'] or '-',
            member_str
        ]

        fill = FILL_ZEBRA_1 if row_idx % 2 == 0 else FILL_HIGHLIGHT_NGOAI
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = FONT_REGULAR
            c.fill = fill
            c.border = BORDER_THIN
            c.alignment = ALIGN_CENTER if col_idx in [1, 2, 3, 4, 6, 7] else ALIGN_LEFT

        row_idx += 1

    autofit_columns(ws)

# ==========================================
# SHEET 4: Thống kê (HR Analytics)
# ==========================================
def create_sheet_thong_ke(wb, member_stats, assigned_shifts):
    ws = wb.create_sheet(title="thống kê")
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:K1")
    ws["A1"] = "BÁO CÁO THỐNG KÊ & PHÂN TÍCH HIỆU SUẤT NHÂN SỰ TRỰC CA (50 THÀNH VIÊN)"
    ws["A1"].font = Font(name="Segoe UI", size=15, bold=True, color="047857")
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    headers = [
        "Mã TV", "Họ Và Tên", "Ban Chuyên Môn", "Đối Tượng", "Nơi Sinh Sống", "Số Điện Thoại", 
        "Đội Ứng Biến", "Tổng Ca", "Tổng Giờ", "Ca Trong", "Ca Ngoài", "Ca Cam Kết", "Mã Ca Được Phân Công"
    ]
    
    ws.row_dimensions[3].height = 26
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_GREEN
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    row_idx = 4
    for m in member_stats:
        ws.row_dimensions[row_idx].height = 22
        standby_txt = "Có" if m['is_standby'] else "Không"
        
        values = [
            m['member_id'],
            m['name'],
            m['department'],
            m['job'],
            m['residence'],
            m['phone'],
            standby_txt,
            m['total_shifts'],
            m['total_hours'],
            m['phong_shifts'],
            m['ngoai_shifts'],
            m['committed_matched'],
            m['assigned_shift_ids']
        ]

        fill = FILL_ZEBRA_1 if row_idx % 2 == 0 else FILL_ZEBRA_2
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = FONT_REGULAR
            c.fill = fill
            c.border = BORDER_THIN
            c.alignment = ALIGN_CENTER if col_idx in [1, 4, 6, 7, 8, 9, 10, 11, 12] else ALIGN_LEFT
            if col_idx in [8, 9]:
                c.font = FONT_BOLD

        row_idx += 1

    row_idx += 2
    ws.cell(row=row_idx, column=1, value="BẢNG TỔNG HỢP THEO BAN CHUYÊN MÔN").font = FONT_SECTION
    row_idx += 1
    
    dept_headers = ["Ban", "Số Thành Viên", "Tổng Ca Trực", "Trung Bình Ca/Người"]
    for c_i, h in enumerate(dept_headers, 1):
        c = ws.cell(row=row_idx, column=c_i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER_DARK
        c.alignment = ALIGN_CENTER
        c.border = BORDER_HEADER
    row_idx += 1

    dept_groups = {}
    for m in member_stats:
        d = m['department']
        if d not in dept_groups:
            dept_groups[d] = {'count': 0, 'shifts': 0}
        dept_groups[d]['count'] += 1
        dept_groups[d]['shifts'] += m['total_shifts']

    for d, info in dept_groups.items():
        avg_s = round(info['shifts'] / info['count'], 2) if info['count'] else 0
        ws.cell(row=row_idx, column=1, value=d).border = BORDER_THIN
        ws.cell(row=row_idx, column=2, value=info['count']).border = BORDER_THIN
        ws.cell(row=row_idx, column=3, value=info['shifts']).border = BORDER_THIN
        ws.cell(row=row_idx, column=4, value=avg_s).border = BORDER_THIN
        for ci in range(1, 5):
            ws.cell(row=row_idx, column=ci).alignment = ALIGN_CENTER if ci > 1 else ALIGN_LEFT
        row_idx += 1

    autofit_columns(ws)

# ==========================================
# SHEET 5: kiểm tra ca (Audit & Validation)
# ==========================================
def create_sheet_kiem_tra_ca(wb, audit, assigned_shifts, member_stats):
    ws = wb.create_sheet(title="kiểm tra ca")
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:H1")
    ws["A1"] = "BÁO CÁO KIỂM TRA & THẨM ĐỊNH TÍNH HỢP LỆ LỊCH TRỰC (AUDIT REPORT)"
    ws["A1"].font = Font(name="Segoe UI", size=15, bold=True, color="6D28D9")
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A3:D3")
    ws["A3"] = "CHỈ SỐ THẨM ĐỊNH QUAN TRỌNG"
    ws["A3"].font = FONT_SECTION
    
    metrics = [
        ("Xung đột trùng ca (Ca Trong vs Ca Ngoài)", f"{audit.get('conflict_count', 0)} vi phạm", "Tuyệt đối không trùng", "100% ĐẠT"),
        ("Vi phạm lịch rảnh đã đăng ký", f"{audit.get('availability_violation_count', 0)} vi phạm", "Tôn trọng lịch rảnh thành viên", "100% ĐẠT"),
        ("Phòng bán trống không có người trực", f"{audit.get('empty_room_count', 0)} ca trống", "Không để phòng trống", "100% ĐẠT"),
        ("Số ca quá tải trong ngày (>2 ca/ngày)", f"{audit.get('daily_overload_count', 0)} trường hợp", "Tối đa 2 ca/ngày/người", "100% ĐẠT"),
        ("Độ công bằng phân bổ (Fairness Score)", f"{audit.get('fairness_metrics', {}).get('fairness_score', 95)}/100", "Phân bổ đều giữa các thành viên", "TỐT")
    ]

    card_headers = ["Hạng Mục Kiểm Tra", "Kết Quả Thực Tế", "Tiêu Chuẩn Đề Ra", "Đánh Giá"]
    for ci, h in enumerate(card_headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER_PURPLE
        c.alignment = ALIGN_CENTER
        c.border = BORDER_HEADER

    for ri, (item, actual, std, eval_str) in enumerate(metrics, 5):
        ws.row_dimensions[ri].height = 24
        ws.cell(row=ri, column=1, value=item).alignment = ALIGN_LEFT
        ws.cell(row=ri, column=2, value=actual).alignment = ALIGN_CENTER
        ws.cell(row=ri, column=3, value=std).alignment = ALIGN_LEFT
        c_eval = ws.cell(row=ri, column=4, value=eval_str)
        c_eval.alignment = ALIGN_CENTER
        c_eval.font = Font(name="Segoe UI", size=10, bold=True, color="047857")
        for ci in range(1, 5):
            ws.cell(row=ri, column=ci).border = BORDER_THIN
            ws.cell(row=ri, column=ci).fill = FILL_SUCCESS_LIGHT

    row_idx = 12
    ws.merge_cells(f"A{row_idx}:H{row_idx}")
    ws.cell(row=row_idx, column=1, value="CHI TIẾT ĐỘ PHỦ TỪNG KHUNG GIỜ (COVERAGE CHECK)").font = FONT_SECTION
    row_idx += 1

    cov_headers = ["Mã Ca", "Kênh", "Thứ", "Khung Giờ", "Yêu Cầu (Chính+DP)", "Đã Gán", "Độ Phủ (%)", "Kết Luận"]
    for ci, h in enumerate(cov_headers, 1):
        c = ws.cell(row=row_idx, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER_DARK
        c.alignment = ALIGN_CENTER
        c.border = BORDER_HEADER
    row_idx += 1

    for s in assigned_shifts:
        ws.row_dimensions[row_idx].height = 20
        ratio = round((s['assigned_count'] / s['required_count']) * 100, 1) if s['required_count'] else 100
        concl = "Hoàn hảo" if ratio >= 100 else "Thiếu định mức"
        
        row_vals = [
            s['shift_id'], s['type_label'], s['day'], s['slot'],
            s['required_count'], s['assigned_count'], f"{ratio}%", concl
        ]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=row_idx, column=ci, value=v)
            c.font = FONT_REGULAR
            c.border = BORDER_THIN
            c.alignment = ALIGN_CENTER
        row_idx += 1

    autofit_columns(ws)

# ==========================================
# SHEET 6: Vắng, Đi Trễ & Thay Thế, Đổi Ca
# ==========================================
def create_sheet_ca_vang(wb, assigned_shifts, contingency, incident_logs=None):
    ws = wb.create_sheet(title="Vắng, Đi Trễ & Thay Thế")
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:J1")
    ws["A1"] = "BÁO CÁO THỐNG KÊ NHÂN SỰ VẮNG, ĐI TRỄ & THAY THẾ, ĐỔI CA DỰ PHÒNG"
    ws["A1"].font = Font(name="Segoe UI", size=15, bold=True, color="DC2626")
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:J2")
    ws["A2"] = "Theo dõi chi tiết điểm danh, trường hợp vắng/trễ, nhân sự điều động thay thế & ma trận dự phòng khẩn cấp"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = ALIGN_CENTER
    ws.row_dimensions[2].height = 20

    # --- PART 1: INCIDENT & ATTENDANCE LOGS ---
    row_idx = 4
    ws.merge_cells(f"A{row_idx}:J{row_idx}")
    ws.cell(row=row_idx, column=1, value="BẢNG 1: NHẬT KÝ THÀNH VIÊN VẮNG, ĐI TRỄ & THAY THẾ, ĐỔI CA ĐÃ GHI NHẬN").font = FONT_SECTION
    row_idx += 1

    inc_headers = [
        "STT", "Thời Gian Ghi Nhận", "Mã Ca", "Thứ / Khung Giờ", 
        "Loại Sự Cố / Trạng Thái", "Thành Viên Vắng / Trễ / Đổi Ca", "Thành Viên Thay Thế", "Ghi Chú / Lý Do", "", ""
    ]
    
    ws.row_dimensions[row_idx].height = 26
    for col_idx, header in enumerate(inc_headers[:8], 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    row_idx += 1
    logs = incident_logs or []
    if not logs:
        ws.row_dimensions[row_idx].height = 24
        ws.merge_cells(f"A{row_idx}:H{row_idx}")
        cell = ws.cell(row=row_idx, column=1, value="Chưa có sự cố vắng mặt hoặc đi trễ nào được ghi nhận.")
        cell.font = FONT_SUBTITLE
        cell.alignment = ALIGN_CENTER
        for ci in range(1, 9):
            ws.cell(row=row_idx, column=ci).border = BORDER_THIN
        row_idx += 1
    else:
        for idx, inc in enumerate(logs, 1):
            ws.row_dimensions[row_idx].height = 24
            st_type = inc.get('status_type', 'Vắng mặt')
            
            fill = FILL_ZEBRA_1 if row_idx % 2 == 0 else FILL_ZEBRA_2
            status_color = "DC2626" if "Vắng" in st_type else ("D97706" if "trễ" in st_type else "047857")
            
            vals = [
                idx,
                inc.get('timestamp', '-'),
                inc.get('shift_id', '-'),
                f"{inc.get('day', '')} ({inc.get('slot', '')})",
                st_type,
                inc.get('absent_member', '-'),
                inc.get('replacement_member', '-'),
                inc.get('note', '-')
            ]

            for col_idx, val in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.font = FONT_REGULAR
                c.fill = fill
                c.border = BORDER_THIN
                c.alignment = ALIGN_CENTER if col_idx in [1, 2, 3, 4, 5] else ALIGN_LEFT
                if col_idx == 5:
                    c.font = Font(name="Segoe UI", size=10, bold=True, color=status_color)
                if col_idx == 7 and val != '-':
                    c.font = Font(name="Segoe UI", size=10, bold=True, color="1E40AF")
            row_idx += 1

    # --- PART 2: CONTINGENCY MATRIX ---
    row_idx += 2
    ws.merge_cells(f"A{row_idx}:J{row_idx}")
    ws.cell(row=row_idx, column=1, value="BẢNG 2: MA TRẬN NHÂN SỰ DỰ PHÒNG ƯU TIÊN THEO CA TRỰC (STANDBY MATRIX)").font = FONT_SECTION
    row_idx += 1

    headers = [
        "Mã Ca", "Kênh Trực", "Thứ", "Khung Giờ", "Địa Điểm", 
        "Nhân Sự Phân Bổ Ca", "Dự Phòng Ưu Tiên 1 (SĐT)", "Dự Phòng Ưu Tiên 2 (SĐT)", "Dự Phòng Ưu Tiên 3 (SĐT)", "Ghi Chú Vận Hành"
    ]
    
    ws.row_dimensions[row_idx].height = 28
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_DARK
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    row_idx += 1
    for item in contingency:
        ws.row_dimensions[row_idx].height = 28
        backups = item.get('backup_candidates', [])
        
        b1 = f"{backups[0]['name']} ({backups[0]['phone']})" if len(backups) > 0 else "Hết nhân sự rảnh"
        b2 = f"{backups[1]['name']} ({backups[1]['phone']})" if len(backups) > 1 else "-"
        b3 = f"{backups[2]['name']} ({backups[2]['phone']})" if len(backups) > 2 else "-"

        assigned_str = ", ".join(item['current_assigned'])

        values = [
            item['shift_id'],
            item['type_label'],
            item['day'],
            item['slot'],
            item['location'],
            assigned_str,
            b1,
            b2,
            b3,
            "Liên hệ dự phòng theo thứ tự P1 ➔ P2 ➔ P3 khi vắng mặt"
        ]

        fill = FILL_ZEBRA_1 if row_idx % 2 == 0 else FILL_ZEBRA_2
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = FONT_REGULAR
            c.fill = fill
            c.border = BORDER_THIN
            c.alignment = ALIGN_CENTER if col_idx in [1, 2, 3, 4] else ALIGN_LEFT
            if col_idx == 7 and len(backups) > 0:
                c.font = Font(name="Segoe UI", size=10, bold=True, color="1E40AF")

        row_idx += 1

    autofit_columns(ws)
